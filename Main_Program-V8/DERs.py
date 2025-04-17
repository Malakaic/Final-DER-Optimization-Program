import csv
import os
import tkinter as tk
import requests
from tkinter import filedialog, messagebox, ttk
import config
import Solar_PV_csv_save
import Location_Input
from openpyxl import Workbook, load_workbook


class Der_menu_page (tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        
        
    def create_der_section(self, frame):

            der_frame = ttk.LabelFrame(frame, text="Select DER Resources")
            der_frame.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)

            # Define a dictionary to store the states of the checkboxes
            checkbox_states = {}

            # Define the options for the checkboxes
            options = ["PV", "Wind", "Battery", "Inverter"]
            
            # Add input boxes for Max Turbines and Max PVs
            max_turbines_label = tk.Label(der_frame, text="Max Turbines:", anchor="w")
            max_turbines_label.pack(anchor="w", padx=5, pady=2)
            max_turbines_entry = tk.Entry(der_frame)
            max_turbines_entry.pack(anchor="w", padx=5, pady=2)

            max_pvs_label = tk.Label(der_frame, text="Max PVs:", anchor="w")
            max_pvs_label.pack(anchor="w", padx=5, pady=2)
            max_pvs_entry = tk.Entry(der_frame)
            max_pvs_entry.pack(anchor="w", padx=5, pady=2)

            # Function to display the selected options
            def show_selected():
                # Save the max turbines and max PVs values to config variables
                config.turbine_max = max_turbines_entry.get()
                config.PV_max = max_pvs_entry.get()
                print(f"Max Turbines: {config.turbine_max}, Max PVs: {config.PV_max}")

                # First, remove any previously created frames
                for child in frame.winfo_children():
                    if isinstance(child, ttk.LabelFrame) and child.cget("text") in options:
                        child.destroy()

                selected = [option for option, var in checkbox_states.items() if var.get()]

                # Loop through each selected option and create a new frame for each
                for idx, option in enumerate(selected):
                    # Calculate row and column for 2x2 grid
                    row, col = divmod(idx, 2)

                    # Create a new frame for each selected option
                    result_frame = ttk.LabelFrame(frame, text=option)
                    result_frame.grid(row=row, column=1 + col, sticky="nsew", padx=5, pady=5)  # Shift to the right column

                    if option == "PV":
                        pv_frame = ttk.Frame(result_frame)  # A container for better layout
                        pv_frame.pack(pady=10, padx=10, fill="x")

                        # Text box for Name
                        pv_name_label = tk.Label(pv_frame, text="Name:", anchor="w")
                        pv_name_label.grid(row=1, column=0, sticky="w", padx=5)
                        pv_name_entry = tk.Entry(pv_frame)
                        pv_name_entry.grid(row=1, column=1, sticky="e", padx=5)

                        # Text box for Size
                        pv_size_label = tk.Label(pv_frame, text="Size (kW-DC):", anchor="w")
                        pv_size_label.grid(row=2, column=0, sticky="w", padx=5)
                        pv_size_entry = tk.Entry(pv_frame)
                        pv_size_entry.grid(row=2, column=1, sticky="e", padx=5)


                        # Text box for Lifespan
                        pv_lifespan_label = tk.Label(pv_frame, text="Lifespan (years):", anchor="w")
                        pv_lifespan_label.grid(row=3, column=0, sticky="w", padx=5)
                        pv_lifespan_entry = tk.Entry(pv_frame)
                        pv_lifespan_entry.grid(row=3, column=1, sticky="e", padx=5)

                        
                        # Text box for efficiency
                        pv_efficiency_label = tk.Label(pv_frame, text="Efficiency (%):", anchor="w")
                        pv_efficiency_label.grid(row=4, column=0, sticky="w", padx=5)
                        pv_efficiency_entry = tk.Entry(pv_frame)
                        pv_efficiency_entry.grid(row=4, column=1, sticky="e", padx=5)

                        # Dropdown for Module type
                        module_type_label = tk.Label(pv_frame, text="Module Type:", anchor="w")
                        module_type_label.grid(row=5, column=0, sticky="w", padx=5)
                    
                        
                        module_types = ["Monocrystalline", "Polycrystalline", "Thin-Film"]
                        module_type_var = tk.StringVar()
                        module_type_dropdown = ttk.Combobox(pv_frame, textvariable=module_type_var, values=module_types, state="readonly")
                        module_type_dropdown.grid(row=5, column=1, sticky="e", padx=5)
                        module_type_dropdown.current(0)  # Set default selection
                        

                        # Text box for cost
                        pv_cost_label = tk.Label(pv_frame, text="Cost ($/kW-DC):", anchor="w")
                        pv_cost_label.grid(row=6, column=0, sticky="w", padx=5)
                        pv_cost_entry = tk.Entry(pv_frame)
                        pv_cost_entry.grid(row=6, column=1, sticky="e", padx=5)


                        def save_pv_data():
                            # Gather data for the current PV configuration

                            pv_data = [
                                pv_name_entry.get(),  # Name
                                pv_size_entry.get(),     # Size
                                pv_lifespan_entry.get(), # Lifespan
                                pv_efficiency_entry.get(),# Efficiency
                                module_type_var.get(),# Module Type
                                pv_cost_entry.get()     # Cost
                            ]

                                            
                            if pv_data[4] == "Monocrystalline":
                                pv_data[4] = int(0)
                            elif pv_data[4] == "Polycrystalline":
                                pv_data[4] = int(1)
                            elif pv_data[4] == "Thin-Film":
                                pv_data[4] = int(2)


                            # Check if a save with the same name already exists
                            existing_key = None
                            for key, value in config.pv_data_dict.items():
                                if value[0] == pv_data[0]:  # Compare the "Name" field
                                    existing_key = key
                        
                            if existing_key is not None:
                                # Overwrite the existing save
                                overwrite = messagebox.askyesno("Overwrite Entry", f"An entry with the name '{pv_data[0]}' already exists. Do you want to overwrite it?")
                                if overwrite:
                                    config.pv_data_dict[existing_key] = pv_data
                                    print(f"PV data with name '{pv_data[0]}' overwritten.")
                                else:
                                    messagebox.showinfo("Change Name", "Please change the name of the PV before saving.")
                            else:
                                # Save Current Data
                                config.pv_data_dict[config.pv_counter] = pv_data
                                print(f"PV data saved: {pv_data}")
                                config.pv_counter += 1
                                
                            # Save data to Excel
                            excel_filename = "pv_data.xlsx"
                            file_exists = os.path.isfile(excel_filename)

                            if file_exists:
                                workbook = load_workbook(excel_filename)
                                sheet = workbook.active
                            else:
                                workbook = Workbook()
                                sheet = workbook.active
                                # Write header row if new file
                                sheet.append(["Name", "Size (kW-DC)", "Lifespan (years)", "Efficiency (%)", "Module Type", "Cost ($/kW-DC)"])

                            # Append data row
                            sheet.append(pv_data)
                            workbook.save(excel_filename)
                            print(f"PV data written to {excel_filename}")
                        
                        def existing_PV_configurations():
                            """Open a popup window to display existing PV configurations and allow selection."""
                            # Create a new window to display existing configurations
                            existing_window = tk.Toplevel(self.parent)
                            existing_window.title("Existing PV Configurations")
                            existing_window.geometry("700x400")

                            # Create a Treeview to display the existing configurations
                            tree = ttk.Treeview(existing_window, columns=("Name", "Size", "Lifespan", "Efficiency", "Module Type", "Cost"), show="headings")
                            tree.heading("Name", text="Name")
                            tree.heading("Size", text="Size (kW-DC)")
                            tree.heading("Lifespan", text="Lifespan (years)")
                            tree.heading("Efficiency", text="Efficiency (%)")
                            tree.heading("Module Type", text="Module Type")
                            tree.heading("Cost", text="Cost ($/kW-DC)")
                            tree.pack(fill="both", expand=True, padx=10, pady=10)

                            # Define column widths
                            tree.column("Name", width=100, anchor="center")
                            tree.column("Size", width=100, anchor="center")
                            tree.column("Lifespan", width=120, anchor="center")
                            tree.column("Efficiency", width=120, anchor="center")
                            tree.column("Module Type", width=150, anchor="center")
                            tree.column("Cost", width=120, anchor="center")

                            # Insert existing configurations into the Treeview
                            for key, value in config.pv_data_existing_configurations.items():
                                tree.insert("", "end", values=value)

                            

                            def select_pv_configuration():
                                """Handle the selection of a configuration."""
                                selected_item = tree.selection()
                                if selected_item:
                                    # Get the selected item's values
                                    selected_values = list(tree.item(selected_item, "values"))  # Convert tuple to list
                                    print(f"Selected PV Configuration: {selected_values}")

                                    # Populate the PV input fields with the selected values
                                    pv_name_entry.delete(0, tk.END)
                                    pv_name_entry.insert(0, selected_values[0])  # Name
                                    pv_size_entry.delete(0, tk.END)
                                    pv_size_entry.insert(0, selected_values[1])  # Size
                                    pv_lifespan_entry.delete(0, tk.END)
                                    pv_lifespan_entry.insert(0, selected_values[2])  # Lifespan
                                    pv_efficiency_entry.delete(0, tk.END)
                                    pv_efficiency_entry.insert(0, selected_values[3])  # Efficiency
                                    module_type_var.set(selected_values[4])  # Module Type
                                    pv_cost_entry.delete(0, tk.END)
                                    pv_cost_entry.insert(0, selected_values[5])  # Cost

                                    
                                    if selected_values[4] == "Monocrystalline":
                                        selected_values[4] = int(0)
                                    elif selected_values[4] == "Polycrystalline":
                                        selected_values[4] = int(1)
                                    elif selected_values[4] == "Thin-Film":
                                        selected_values[4] = int(2)

                                    # Save the selected configuration to the global dictionary
                                    config.pv_data_dict[config.pv_counter] = list(selected_values)
                                    config.pv_counter += 1
                                    print(f"Selected configuration saved: {selected_values}")

                                    # Close the popup window
                                    existing_window.destroy()
                                else:
                                    messagebox.showwarning("Selection Error", "Please select a configuration.")
                                # Add buttons to the popup window
                            select_button = tk.Button(existing_window, text="Select Configuration", command=select_pv_configuration)
                            select_button.pack(pady=10)

                            close_button = tk.Button(existing_window, text="Close", command=existing_window.destroy)
                            close_button.pack(pady=10)

                        pv_configurations_button = tk.Button(pv_frame, text="Exisiting PV Configurations", command=existing_PV_configurations)
                        pv_configurations_button.grid(row=0, column=0, columnspan=2, pady=10)


                        pv_save_button = tk.Button(pv_frame, text="Save", command=save_pv_data)
                        pv_save_button.grid(row=8, column=0, columnspan=2, pady=10)
                        

                    if option == "Wind":
                        # A container for the layout of all Wind-specific inputs
                        wind_frame = ttk.Frame(result_frame)
                        wind_frame.pack(pady=10, padx=10, fill="x")

                        # "Name" label and entry
                        wind_name_label = tk.Label(wind_frame, text="Name:", anchor="w")
                        wind_name_label.grid(row=1, column=0, sticky="w", padx=5)
                        wind_name_entry = tk.Entry(wind_frame)
                        wind_name_entry.grid(row=1, column=1, sticky="e", padx=5)

                        # "Size" label and entry
                        wind_size_label = tk.Label(wind_frame, text="Size (kW AC):", anchor="w")
                        wind_size_label.grid(row=2, column=0, sticky="w", padx=5)
                        wind_size_entry = tk.Entry(wind_frame)
                        wind_size_entry.grid(row=2, column=1, sticky="e", padx=5)

                        # "Lifespan" label and entry
                        wind_lifespan_label = tk.Label(wind_frame, text="Lifespan (years):", anchor="w")
                        wind_lifespan_label.grid(row=3, column=0, sticky="w", padx=5)
                        wind_lifespan_entry = tk.Entry(wind_frame)
                        wind_lifespan_entry.grid(row=3, column=1, sticky="e", padx=5)


                        # "Efficiency" label and entry
                        wind_efficiency_label = tk.Label(wind_frame, text="Efficiency (%):", anchor="w")
                        wind_efficiency_label.grid(row=4, column=0, sticky="w", padx=5)
                        wind_efficiency_entry = tk.Entry(wind_frame)
                        wind_efficiency_entry.grid(row=4, column=1, sticky="e", padx=5)

                        # "Hub Height" label and entry
                        hub_height_label = tk.Label(wind_frame, text="Hub Height (meters):", anchor="w")
                        hub_height_label.grid(row=5, column=0, sticky="w", padx=5)
                        hub_height_entry = tk.Entry(wind_frame)
                        hub_height_entry.grid(row=5, column=1, sticky="e", padx=5)

                        # "Rotor Diameter" label and entry
                        rotor_diameter_label = tk.Label(wind_frame, text="Rotor Diameter (meters):", anchor="w")
                        rotor_diameter_label.grid(row=6, column=0, sticky="w", padx=5)
                        rotor_diameter_entry = tk.Entry(wind_frame)
                        rotor_diameter_entry.grid(row=6, column=1, sticky="e", padx=5)

                        # "Cost" label and entry
                        turbine_cost_label = tk.Label(wind_frame, text="Cost ($/kWh):", anchor="w")
                        turbine_cost_label.grid(row=7, column=0, sticky="w", padx=5)
                        turbine_cost_entry = tk.Entry(wind_frame)
                        turbine_cost_entry.grid(row=7, column=1, sticky="e", padx=5)


                        # Save Button for Wind
                        def save_wind_data():
                            # Create a list with the data for the current turbine
                            wind_data = [
                                wind_name_entry.get(),   # Name
                                wind_size_entry.get(),        # Size
                                wind_efficiency_entry.get(),   # Efficiency
                                wind_lifespan_entry.get(),    # Lifespan
                                hub_height_entry.get(),  # Hub Height
                                rotor_diameter_entry.get(),  # Rotor Diameter
                                turbine_cost_entry.get()  # Cost
                            ]

                            # Check if a save with the same name already exists
                            existing_key = None
                            for key, value in config.wind_data_dict.items():
                                if value[0] == wind_data[0]:  # Compare the "Name" field
                                    existing_key = key

                            if existing_key is not None:
                                # Overwrite the existing save
                                overwrite = messagebox.askyesno("Overwrite Entry", f"An entry with the name '{wind_data[0]}' already exists. Do you want to overwrite it?")
                                if overwrite:
                                    config.wind_data_dict[existing_key] = wind_data
                                    print(f"Wind data with name '{wind_data[0]}' overwritten.")
                                else:
                                    messagebox.showinfo("Change Name", "Please change the name of the wind turbine before saving.")
                            else:
                                # Save Current Data
                                config.wind_data_dict[config.wind_counter] = wind_data
                                print(f"Wind data saved: '{wind_data}'.")
                                config.wind_counter += 1

                        def existing_wind_configurations():
                            """Open a popup window to display existing PV configurations and allow selection."""
                            # Create a new window to display existing configurations
                            existing_window = tk.Toplevel(self.parent)
                            existing_window.title("Existing Wind Configurations")
                            existing_window.geometry("700x400")

                            # Create a Treeview to display the existing configurations
                            
                            tree = ttk.Treeview(existing_window, columns=("Name", "Size", "Lifespan", "Efficiency", "Hub Height", "Rotor Diameter", "Cost"), show="headings")
                            tree.heading("Name", text="Name")
                            tree.heading("Size", text="Size (kW)")
                            tree.heading("Lifespan", text="Lifespan (years)")
                            tree.heading("Efficiency", text="Efficiency (%)")
                            tree.heading("Hub Height", text="Hub Height (m)")
                            tree.heading("Rotor Diameter", text="Rotor Diameter (m)")
                            tree.heading("Cost", text="Cost ($/kW-AC)")
                            tree.pack(fill="both", expand=True, padx=10, pady=10)

                            # Define column widths
                            tree.column("Name", width=100, anchor="center")
                            tree.column("Size", width=100, anchor="center")
                            tree.column("Lifespan", width=120, anchor="center")
                            tree.column("Efficiency", width=120, anchor="center")
                            tree.column("Hub Height", width=150, anchor="center")
                            tree.column("Rotor Diameter", width=150, anchor="center")
                            tree.column("Cost", width=120, anchor="center")

                            # Insert existing configurations into the Treeview
                            for key, value in config.wind_data_existing_configurations.items():
                                tree.insert("", "end", values=value)

                            

                            def select_wind_configuration():
                                """Handle the selection of a configuration."""
                                selected_item = tree.selection()
                                if selected_item:
                                    # Get the selected item's values
                                    selected_values = list(tree.item(selected_item, "values"))  # Convert tuple to list
                                    print(f"Selected Wind Turbine Configuration: {selected_values}")

                                    # Populate the PV input fields with the selected values
                                    wind_name_entry.delete(0, tk.END)
                                    wind_name_entry.insert(0, selected_values[0])  # Name
                                    wind_size_entry.delete(0, tk.END)
                                    wind_size_entry.insert(0, selected_values[1])  # Size
                                    wind_lifespan_entry.delete(0, tk.END)
                                    wind_lifespan_entry.insert(0, selected_values[2])  # Lifespan
                                    wind_efficiency_entry.delete(0, tk.END)
                                    wind_efficiency_entry.insert(0, selected_values[3])  # Efficiency
                                    hub_height_entry.delete(0, tk.END)  # Hub Height
                                    hub_height_entry.insert(0, selected_values[4])  # Hub Height
                                    rotor_diameter_entry.delete(0, tk.END)  # Rotor Diameter
                                    rotor_diameter_entry.insert(0, selected_values[5])  # Rotor Diameter
                                    turbine_cost_entry.delete(0, tk.END) 
                                    turbine_cost_entry.insert(0, selected_values[6])  # Cost

                                    # Save the selected configuration to the global dictionary
                                    config.wind_data_dict[config.wind_counter] = list(selected_values)
                                    config.wind_counter += 1
                                    print(f"Selected configuration saved: {selected_values}")

                                    # Close the popup window
                                    existing_window.destroy()
                                else:
                                    messagebox.showwarning("Selection Error", "Please select a configuration.")
                                # Add buttons to the popup window
                            select_button = tk.Button(existing_window, text="Select Configuration", command=select_wind_configuration)
                            select_button.pack(pady=10)

                            close_button = tk.Button(existing_window, text="Close", command=existing_window.destroy)
                            close_button.pack(pady=10)

                        wind_configurations_button = tk.Button(wind_frame, text="Existing Wind Turbine Configurations", command=existing_wind_configurations)
                        wind_configurations_button.grid(row=0, column=0, columnspan=2, pady=10)

                        wind_save_button = tk.Button(wind_frame, text="Save", command=save_wind_data)
                        wind_save_button.grid(row=8, column=0, columnspan=2, pady=10)




                    if option == "Battery":
                        # A container for the layout of all Battery-specific inputs
                        battery_frame = ttk.Frame(result_frame)
                        battery_frame.pack(pady=10, padx=10, fill="x")

                        # "Name" label and entry
                        battery_name_label = tk.Label(battery_frame, text="Name:", anchor="w")
                        battery_name_label.grid(row=0, column=0, sticky="w", padx=5)
                        battery_name_entry = tk.Entry(battery_frame)
                        battery_name_entry.grid(row=0, column=1, sticky="e", padx=5)

                        # "Energy capacity cost ($/kWh)" label and entry
                        energy_cost_label = tk.Label(battery_frame, text="Energy capacity cost ($/kWh):", anchor="w")
                        energy_cost_label.grid(row=1, column=0, sticky="w", padx=5)
                        energy_cost_entry = tk.Entry(battery_frame)
                        energy_cost_entry.grid(row=1, column=1, sticky="e", padx=5)

                        # "Power capacity cost ($/kW)" label and entry
                        power_cost_label = tk.Label(battery_frame, text="Power capacity cost ($/kW):", anchor="w")
                        power_cost_label.grid(row=2, column=0, sticky="w", padx=5)
                        power_cost_entry = tk.Entry(battery_frame)
                        power_cost_entry.grid(row=2, column=1, sticky="e", padx=5)

                        # "Allow grid to charge battery" checkbox
                        grid_charge_var = tk.BooleanVar()
                        grid_charge_checkbox = tk.Checkbutton(
                            battery_frame, text="Allow grid to charge battery", variable=grid_charge_var, anchor="w"
                        )
                        grid_charge_checkbox.grid(row=3, column=0, columnspan=2, sticky="w", padx=5)

                        # "Minimum energy capacity (kWh)" label and entry
                        min_energy_label = tk.Label(battery_frame, text="Minimum energy capacity (kWh):", anchor="w")
                        min_energy_label.grid(row=4, column=0, sticky="w", padx=5)
                        min_energy_entry = tk.Entry(battery_frame)
                        min_energy_entry.grid(row=4, column=1, sticky="e", padx=5)

                        # "Maximum energy capacity (kWh)" label and entry
                        max_energy_label = tk.Label(battery_frame, text="Maximum energy capacity (kWh):", anchor="w")
                        max_energy_label.grid(row=5, column=0, sticky="w", padx=5)
                        max_energy_entry = tk.Entry(battery_frame)
                        max_energy_entry.grid(row=5, column=1, sticky="e", padx=5)

                        # Save Button for Battery
                        def save_battery_data():

                            # Create a list with the data for the current battery
                            battery_data = [
                                battery_name_entry.get(),
                                energy_cost_entry.get(),
                                power_cost_entry.get(),
                                grid_charge_var.get(),
                                min_energy_entry.get(),
                                max_energy_entry.get()
                            ]

                            # Check if a save with the same name already exists
                            existing_key = None
                            for key, value in config.battery_data_dict.items():
                                if value[0] == battery_data[0]:
                                    existing_key = key

                            if existing_key is not None:
                                # Overwrite the existing save
                                # Show a message box indicating the overwrite
                                overwrite = messagebox.askyesno("Overwrite Entry", f"An entry with the name '{battery_data[0]}' already exists. Do you want to overwrite it?")
                                if overwrite:
                                    # Overwrite the existing save
                                    config.battery_data_dict[existing_key] = battery_data
                                    print(f"Battery data with name '{battery_data[0]}' overwritten.")
                                else:
                                    # Prompt user to change the name
                                    messagebox.showinfo("Change Name", "Please change the name of the battery before saving.")
                            else:
                                #Save Current Data
                                config.battery_data_dict[config.battery_counter] = battery_data
                                print(f"Battery data saved: '{battery_data}'.")
                                # Increment the counter for the next save
                                config.battery_counter += 1

                        battery_save_button = tk.Button(battery_frame, text="Save", command=save_battery_data)
                        battery_save_button.grid(row=6, column=0, columnspan=2, pady=10)



            # Create checkboxes dynamically inside der_frame
            for option in options:
                var = tk.BooleanVar()  # Boolean variable for each checkbox
                checkbox_states[option] = var
                checkbox = tk.Checkbutton(der_frame, text=option, variable=var)
                checkbox.pack(anchor="w")  # Align checkboxes to the left

            # Create a button to confirm the selection inside der_frame
            confirm_button = tk.Button(der_frame, text="Confirm Selection", command=show_selected)
            confirm_button.pack(pady=10)

            # Label to display the selected options inside der_frame
            selection_label = tk.Label(der_frame, text="")
            selection_label.pack(pady=10)



    def create_der_input_fields(self, frame):
        """Creates input fields for DER entries."""
        der_input_frame = ttk.Frame(frame)
        der_input_frame.grid(row=0, column=0, padx=5, pady=5)

        tk.Label(der_input_frame, text="Name:").grid(row=0, column=0)
        self.der_name_entry = tk.Entry(der_input_frame)
        self.der_name_entry.grid(row=0, column=1)

        tk.Label(der_input_frame, text="Type:").grid(row=0, column=2)
        self.der_type_entry = ttk.Combobox(der_input_frame, values=["PV", "Wind Turbine", "Battery", "Converter"])
        self.der_type_entry.grid(row=0, column=3)

        tk.Label(der_input_frame, text="Cost ($/kWh):").grid(row=1, column=0)
        self.der_cost_entry = tk.Entry(der_input_frame)
        self.der_cost_entry.grid(row=1, column=1)

        tk.Label(der_input_frame, text="Capacity (kW):").grid(row=1, column=2)
        self.der_capacity_entry = tk.Entry(der_input_frame)
        self.der_capacity_entry.grid(row=1, column=3)

        tk.Label(der_input_frame, text="Efficiency (%):").grid(row=2, column=0)
        self.der_efficiency_entry = tk.Entry(der_input_frame)
        self.der_efficiency_entry.grid(row=2, column=1)

        self.add_der_button = tk.Button(der_input_frame, text="Add DER", command=self.add_der)
        self.add_der_button.grid(row=2, column=2, columnspan=2)

    def add_der(self):
        """Add a DER entry to the Treeview."""
        name = self.der_name_entry.get()
        type_ = self.der_type_entry.get()
        cost = self.der_cost_entry.get()
        capacity = self.der_capacity_entry.get()
        efficiency = self.der_efficiency_entry.get()

        if name and type_ and cost and capacity and efficiency:
            self.der_tree.insert("", "end", values=(name, type_, cost, capacity, efficiency))
            self.clear_der_entries()
        else:
            messagebox.showwarning("Input Error", "Please fill in all fields.")

    def remove_selected_der(self):
        """Remove the selected DER entry from the Treeview."""
        selected_item = self.der_tree.selection()
        if selected_item:
            self.der_tree.delete(selected_item)
        else:
            messagebox.showwarning("Selection Error", "Please select a DER entry to remove.")

    def clear_der_entries(self):
        """Clear DER input fields."""
        self.der_name_entry.delete(0, tk.END)
        self.der_type_entry.set('')  # Reset the dropdown
        self.der_cost_entry.delete(0, tk.END)
        self.der_capacity_entry.delete(0, tk.END)
        self.der_efficiency_entry.delete(0, tk.END)